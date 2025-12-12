"""
Excel report generation for WDD Lock Analysis and Oracle Errors.
"""

from collections import Counter


def generate_excel(results: list[dict], output_excel: str, stats: dict, oracle_errors: list[dict] = None):
    """Generate an Excel file with WDD lock results and Oracle errors on separate sheets."""
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter
    except ImportError:
        print("openpyxl not installed. Install with: pip install openpyxl")
        return

    oracle_errors = oracle_errors or []

    if not results and not oracle_errors:
        print("No results to generate Excel.")
        return

    wb = Workbook()

    # Define common styles
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    red_fill = PatternFill(start_color="FFCCCC", end_color="FFCCCC", fill_type="solid")
    yellow_fill = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
    green_fill = PatternFill(start_color="CCFFCC", end_color="CCFFCC", fill_type="solid")
    error_fill = PatternFill(start_color="FFE6E6", end_color="FFE6E6", fill_type="solid")
    red_font = Font(bold=True, color="8B0000")
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    center_align = Alignment(horizontal='center', vertical='center')
    left_align = Alignment(horizontal='left', vertical='center', wrap_text=True)

    # ==================== Sheet 1: WDD Lock Results ====================
    ws_locks = wb.active
    ws_locks.title = "WDD Lock Results"

    if results:
        # Title row
        ws_locks.merge_cells('A1:F1')
        ws_locks['A1'] = "WDD Lock Analysis Report"
        ws_locks['A1'].font = Font(bold=True, size=14)
        ws_locks['A1'].alignment = Alignment(horizontal='center')

        # Stats row
        ws_locks.merge_cells('A2:F2')
        ws_locks['A2'] = f"Total: {stats['total']} | Success: {stats['success']} | Failed: {stats['failed']} | With Delay: {stats['with_delay']}"
        ws_locks['A2'].alignment = Alignment(horizontal='center')

        # Legend row
        ws_locks['A3'] = "Legend:"
        ws_locks['A3'].font = Font(bold=True)
        ws_locks['B3'] = "RED = LOCK FAILED"
        ws_locks['B3'].fill = red_fill
        ws_locks['C3'] = "YELLOW = Time Diff > 0"
        ws_locks['C3'].fill = yellow_fill
        ws_locks['D3'] = "GREEN = Success (no delay)"
        ws_locks['D3'].fill = green_fill

        # Headers (row 5)
        headers = ['File', 'Del ID', 'Wait Start', 'Result Time', 'Time Diff (s)', 'Result']
        for col, header in enumerate(headers, 1):
            cell = ws_locks.cell(row=5, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = center_align
            cell.border = thin_border

        # Data rows
        for row_idx, r in enumerate(results, 6):
            row_data = [
                r['file'],
                r['del_id'],
                r['wait_start'],
                r['result_time'],
                r['time_diff_seconds'],
                r['result']
            ]

            # Determine row fill color
            if r['result'] == 'LOCK FAILED':
                fill = red_fill
            elif r['time_diff_seconds'] > 0:
                fill = yellow_fill
            else:
                fill = green_fill

            for col, value in enumerate(row_data, 1):
                cell = ws_locks.cell(row=row_idx, column=col, value=value)
                cell.fill = fill
                cell.border = thin_border
                cell.alignment = center_align

                # Bold red font for LOCK FAILED in Result column
                if col == 6 and value == 'LOCK FAILED':
                    cell.font = red_font

        # Adjust column widths
        ws_locks.column_dimensions['A'].width = 35
        ws_locks.column_dimensions['B'].width = 15
        ws_locks.column_dimensions['C'].width = 22
        ws_locks.column_dimensions['D'].width = 22
        ws_locks.column_dimensions['E'].width = 15
        ws_locks.column_dimensions['F'].width = 15
    else:
        ws_locks['A1'] = "No WDD Lock Results Found"
        ws_locks['A1'].font = Font(bold=True, size=14)

    # ==================== Sheet 2: Oracle Errors ====================
    ws_errors = wb.create_sheet(title="Oracle Errors")

    # Title row
    ws_errors.merge_cells('A1:G1')
    ws_errors['A1'] = "Oracle Database Errors (excluding ORA-01403)"
    ws_errors['A1'].font = Font(bold=True, size=14)
    ws_errors['A1'].alignment = Alignment(horizontal='center')

    if oracle_errors:
        # Error count row
        ws_errors.merge_cells('A2:G2')
        ws_errors['A2'] = f"Total Errors Found: {len(oracle_errors)}"
        ws_errors['A2'].alignment = Alignment(horizontal='center')
        ws_errors['A2'].font = Font(bold=True, color="8B0000")

        # Headers (row 4)
        error_headers = ['Error Code', 'File', 'Line #', 'Timestamp', 'Context', 'Error Message', 'Full Line']
        for col, header in enumerate(error_headers, 1):
            cell = ws_errors.cell(row=4, column=col, value=header)
            cell.font = header_font
            cell.fill = PatternFill(start_color="DC3545", end_color="DC3545", fill_type="solid")
            cell.alignment = center_align
            cell.border = thin_border

        # Data rows
        for row_idx, err in enumerate(oracle_errors, 5):
            row_data = [
                err['error_code'],
                err['file'],
                err['line_number'],
                err['timestamp'],
                err['context'],
                err['message'],
                err['full_line']
            ]

            for col, value in enumerate(row_data, 1):
                cell = ws_errors.cell(row=row_idx, column=col, value=value)
                cell.fill = error_fill
                cell.border = thin_border
                if col in [6, 7]:  # Message and Full Line columns
                    cell.alignment = left_align
                else:
                    cell.alignment = center_align

                # Bold red font for error code
                if col == 1:
                    cell.font = red_font

        # Adjust column widths
        ws_errors.column_dimensions['A'].width = 12
        ws_errors.column_dimensions['B'].width = 30
        ws_errors.column_dimensions['C'].width = 10
        ws_errors.column_dimensions['D'].width = 20
        ws_errors.column_dimensions['E'].width = 25
        ws_errors.column_dimensions['F'].width = 50
        ws_errors.column_dimensions['G'].width = 80
    else:
        ws_errors['A3'] = "No Oracle errors found (excluding ORA-01403)"
        ws_errors['A3'].font = Font(color="28A745")

    # ==================== Sheet 3: Error Summary ====================
    ws_summary = wb.create_sheet(title="Error Summary")

    # Title
    ws_summary.merge_cells('A1:C1')
    ws_summary['A1'] = "Oracle Error Summary by Error Code"
    ws_summary['A1'].font = Font(bold=True, size=14)
    ws_summary['A1'].alignment = Alignment(horizontal='center')

    if oracle_errors:
        error_counts = Counter(e['error_code'] for e in oracle_errors)

        # Headers
        summary_headers = ['Error Code', 'Count', 'Percentage']
        for col, header in enumerate(summary_headers, 1):
            cell = ws_summary.cell(row=3, column=col, value=header)
            cell.font = header_font
            cell.fill = PatternFill(start_color="6F42C1", end_color="6F42C1", fill_type="solid")
            cell.alignment = center_align
            cell.border = thin_border

        # Data
        total_errors = len(oracle_errors)
        for row_idx, (code, count) in enumerate(error_counts.most_common(), 4):
            percentage = (count / total_errors) * 100

            ws_summary.cell(row=row_idx, column=1, value=code).font = red_font
            ws_summary.cell(row=row_idx, column=1).alignment = center_align
            ws_summary.cell(row=row_idx, column=1).border = thin_border

            ws_summary.cell(row=row_idx, column=2, value=count).alignment = center_align
            ws_summary.cell(row=row_idx, column=2).border = thin_border

            ws_summary.cell(row=row_idx, column=3, value=f"{percentage:.1f}%").alignment = center_align
            ws_summary.cell(row=row_idx, column=3).border = thin_border

        # Column widths
        ws_summary.column_dimensions['A'].width = 15
        ws_summary.column_dimensions['B'].width = 12
        ws_summary.column_dimensions['C'].width = 12

        # Add error descriptions (common Oracle errors)
        error_descriptions = {
            'ORA-00001': 'Unique constraint violated',
            'ORA-00054': 'Resource busy and acquire with NOWAIT specified',
            'ORA-00060': 'Deadlock detected while waiting for resource',
            'ORA-00904': 'Invalid identifier',
            'ORA-00942': 'Table or view does not exist',
            'ORA-01000': 'Maximum open cursors exceeded',
            'ORA-01017': 'Invalid username/password',
            'ORA-01400': 'Cannot insert NULL into column',
            'ORA-01422': 'Exact fetch returns more than requested number of rows',
            'ORA-01427': 'Single-row subquery returns more than one row',
            'ORA-01438': 'Value larger than specified precision',
            'ORA-01476': 'Divisor is equal to zero',
            'ORA-01555': 'Snapshot too old',
            'ORA-01652': 'Unable to extend temp segment',
            'ORA-01722': 'Invalid number',
            'ORA-02049': 'Distributed transaction timeout',
            'ORA-02291': 'Integrity constraint violated - parent key not found',
            'ORA-02292': 'Integrity constraint violated - child record found',
            'ORA-04031': 'Unable to allocate shared memory',
            'ORA-06502': 'PL/SQL: numeric or value error',
            'ORA-06512': 'PL/SQL: at line (stack trace)',
            'ORA-12154': 'TNS: could not resolve connect identifier',
            'ORA-12170': 'TNS: connect timeout occurred',
            'ORA-12541': 'TNS: no listener',
            'ORA-20000': 'User-defined error (RAISE_APPLICATION_ERROR)',
        }

        # Add description column if we have known errors
        ws_summary['E1'] = "Common Error Descriptions"
        ws_summary['E1'].font = Font(bold=True, size=12)
        ws_summary.merge_cells('E1:F1')

        ws_summary.cell(row=3, column=5, value='Error Code').font = header_font
        ws_summary.cell(row=3, column=5).fill = PatternFill(start_color="17A2B8", end_color="17A2B8", fill_type="solid")
        ws_summary.cell(row=3, column=5).alignment = center_align
        ws_summary.cell(row=3, column=5).border = thin_border

        ws_summary.cell(row=3, column=6, value='Description').font = header_font
        ws_summary.cell(row=3, column=6).fill = PatternFill(start_color="17A2B8", end_color="17A2B8", fill_type="solid")
        ws_summary.cell(row=3, column=6).alignment = center_align
        ws_summary.cell(row=3, column=6).border = thin_border

        row = 4
        for code, desc in error_descriptions.items():
            ws_summary.cell(row=row, column=5, value=code).alignment = center_align
            ws_summary.cell(row=row, column=5).border = thin_border
            ws_summary.cell(row=row, column=6, value=desc).alignment = left_align
            ws_summary.cell(row=row, column=6).border = thin_border
            row += 1

        ws_summary.column_dimensions['E'].width = 15
        ws_summary.column_dimensions['F'].width = 50

    else:
        ws_summary['A3'] = "No Oracle errors to summarize"
        ws_summary['A3'].font = Font(color="28A745")

    # Save workbook
    wb.save(output_excel)
    print(f"Excel saved to: {output_excel}")
